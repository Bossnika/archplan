"""
Operatii Batch pe fisiere DWG — Studio Office Kolectiv

Utilitar pentru procesarea automata a mai multor fisiere DWG:
  - Plotare/Export PDF batch
  - Aplicare layere standard
  - Extragere informatii din blocuri cu atribute
  - Redenumire si organizare fisiere

Utilizare:
  python batch-operatii.py --actiune plot --folder ./desene --output ./pdf
  python batch-operatii.py --actiune extrage-blocuri --folder ./desene
  python batch-operatii.py --actiune aplica-layere --tip arhi --folder ./desene

Dependente: pip install pyautocad openpyxl comtypes tqdm
"""

import os
import sys
import argparse
import time
import json
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("pip install openpyxl")

try:
    import win32com.client
    HAS_WIN32 = True
except ImportError:
    HAS_WIN32 = False

# ─────────────────────────────────────────────────────
#  Conectare AutoCAD
# ─────────────────────────────────────────────────────

def get_autocad():
    if not HAS_WIN32:
        raise RuntimeError("win32com nu e disponibil. pip install pywin32")
    acad = win32com.client.Dispatch("AutoCAD.Application")
    acad.Visible = True
    return acad


def deschide_dwg(acad, cale_dwg):
    """Deschide un fisier DWG si returneaza documentul."""
    doc = acad.Documents.Open(str(cale_dwg))
    time.sleep(1.0)  # Asteapta incarcarea
    return doc


# ─────────────────────────────────────────────────────
#  Plotare / Export PDF
# ─────────────────────────────────────────────────────

def export_pdf_batch(folder_dwg, folder_output, format_hartie="A1", orientare="Landscape"):
    """
    Exporta toate DWG-urile din folder in PDF.
    Exporteaza toate layout-urile din fiecare fisier.
    """
    folder_dwg    = Path(folder_dwg)
    folder_output = Path(folder_output)
    folder_output.mkdir(parents=True, exist_ok=True)

    fisiere = list(folder_dwg.glob("*.dwg"))
    if not fisiere:
        print(f"Nu s-au gasit fisiere DWG in: {folder_dwg}")
        return

    print(f"\nExport PDF — {len(fisiere)} fisiere DWG")
    print(f"Output: {folder_output}\n")

    acad = get_autocad()

    rezultate = {"success": [], "errors": []}

    for i, dwg_path in enumerate(fisiere, 1):
        print(f"[{i}/{len(fisiere)}] {dwg_path.name} ...", end=" ")
        try:
            doc = deschide_dwg(acad, dwg_path)
            layouts = doc.Layouts

            for layout in layouts:
                if layout.Name == "Model":
                    continue

                layout.CopyFrom(layout)
                plot_obj = doc.Plot

                # Configurare plot
                plot_cfg = doc.ActiveLayout.PlotConfiguration
                plot_cfg.SetPlotDevice("PDF")
                plot_cfg.SetCanonicalMediaName(format_hartie)
                plot_cfg.SetPlotToFile(True)

                # Nume fisier output
                fisier_pdf = folder_output / f"{dwg_path.stem}_{layout.Name}.pdf"

                plot_obj.PlotToFile(
                    str(fisier_pdf),
                    layout.Name
                )

                print(f"  → {fisier_pdf.name}")
                rezultate["success"].append(str(fisier_pdf))

            doc.Close(False)

        except Exception as e:
            print(f"  EROARE: {e}")
            rezultate["errors"].append({"fisier": str(dwg_path), "eroare": str(e)})

    print(f"\nRezultat: {len(rezultate['success'])} PDF-uri generate, {len(rezultate['errors'])} erori.")
    return rezultate


# ─────────────────────────────────────────────────────
#  Extragere atribute blocuri
# ─────────────────────────────────────────────────────

def extrage_atribute_blocuri(folder_dwg, fisier_output_excel=None):
    """
    Extrage toate atributele din blocurile DWG si
    genereaza un raport Excel cu toate valorile.
    """
    folder_dwg = Path(folder_dwg)
    fisiere    = list(folder_dwg.glob("**/*.dwg"))  # recursiv

    if not fisiere:
        print(f"Nu s-au gasit DWG-uri in: {folder_dwg}")
        return

    print(f"\nExtragere atribute din {len(fisiere)} fisiere DWG...")

    acad = get_autocad()

    toate_datele = []  # Lista de dictionare

    for dwg_path in fisiere:
        print(f"  Procesare: {dwg_path.name}")
        try:
            doc   = deschide_dwg(acad, dwg_path)
            model = doc.ModelSpace

            for entitate in model:
                if entitate.ObjectName == "AcDbBlockReference":
                    try:
                        bloc_name = entitate.Name
                        atribute = {}

                        # Extrage atribute
                        if entitate.HasAttributes:
                            for attr in entitate.GetAttributes():
                                atribute[attr.TagString] = attr.TextString

                        if atribute:
                            rand = {
                                "fisier": dwg_path.name,
                                "bloc":   bloc_name,
                                "layer":  entitate.Layer,
                                "x":      round(entitate.InsertionPoint[0], 3),
                                "y":      round(entitate.InsertionPoint[1], 3),
                            }
                            rand.update(atribute)
                            toate_datele.append(rand)
                    except:
                        pass

            doc.Close(False)
        except Exception as e:
            print(f"    EROARE: {e}")

    if not toate_datele:
        print("Nu s-au gasit blocuri cu atribute.")
        return

    # Genereaza Excel
    wb    = openpyxl.Workbook()
    ws    = wb.active
    ws.title = "Atribute Blocuri"

    # Colecteaza toate cheile (coloanele)
    toate_cheile = set()
    for rand in toate_datele:
        toate_cheile.update(rand.keys())

    coloane_fixe = ["fisier", "bloc", "layer", "x", "y"]
    coloane_attr = sorted(toate_cheile - set(coloane_fixe))
    coloane      = coloane_fixe + coloane_attr

    # Header
    from openpyxl.styles import Font, PatternFill, Alignment
    for col_idx, col_name in enumerate(coloane, 1):
        celula = ws.cell(row=1, column=col_idx, value=col_name.upper())
        celula.fill = PatternFill("solid", fgColor="1F3864")
        celula.font = Font(bold=True, color="FFFFFF")
        celula.alignment = Alignment(horizontal="center")
        ws.column_dimensions[chr(64 + col_idx)].width = max(15, len(col_name) + 2)

    # Date
    for row_idx, rand in enumerate(toate_datele, 2):
        for col_idx, col_name in enumerate(coloane, 1):
            ws.cell(row=row_idx, column=col_idx, value=rand.get(col_name, ""))

    if not fisier_output_excel:
        fisier_output_excel = f"Atribute_Blocuri_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

    wb.save(fisier_output_excel)
    print(f"\n{len(toate_datele)} inregistrari → {fisier_output_excel}")
    return fisier_output_excel


# ─────────────────────────────────────────────────────
#  Aplicare layere standard dintr-un DWG template
# ─────────────────────────────────────────────────────

def aplica_layere_standard(folder_dwg, tip="arhi", template_dwg=None):
    """
    Aplica sistemul standard de layere in toate DWG-urile.
    Executa scriptul LISP de layere in fiecare fisier.
    """
    folder_dwg = Path(folder_dwg)
    fisiere    = list(folder_dwg.glob("*.dwg"))

    if not fisiere:
        print("Nu s-au gasit fisiere DWG.")
        return

    comanda_lisp = "(C:LAYERE-ARHI)" if tip == "arhi" else "(C:LAYERE-URB)"

    print(f"\nAplicare layere {tip.upper()} in {len(fisiere)} fisiere...")
    acad = get_autocad()

    for dwg_path in fisiere:
        print(f"  {dwg_path.name} ...", end=" ")
        try:
            doc = deschide_dwg(acad, dwg_path)
            doc.SendCommand(f"{comanda_lisp}\n")
            time.sleep(0.5)
            doc.Save()
            doc.Close(False)
            print("OK")
        except Exception as e:
            print(f"EROARE: {e}")

    print(f"\nLayere aplicate in {len(fisiere)} fisiere.")


# ─────────────────────────────────────────────────────
#  Audit fisiere DWG
# ─────────────────────────────────────────────────────

def audit_fisiere(folder_dwg, fisier_raport=None):
    """
    Genereaza un raport de audit pentru toate DWG-urile:
    - Dimensiune fisier
    - Numar entitati
    - Layere folosite
    - Blocuri utilizate
    - Data ultima modificare
    """
    folder_dwg = Path(folder_dwg)
    fisiere    = list(folder_dwg.glob("**/*.dwg"))

    print(f"\nAudit {len(fisiere)} fisiere DWG...")

    raport = []

    for dwg_path in fisiere:
        stat = dwg_path.stat()
        info = {
            "fisier":         dwg_path.name,
            "cale":           str(dwg_path.parent),
            "dimensiune_kb":  round(stat.st_size / 1024, 1),
            "ultima_modif":   datetime.fromtimestamp(stat.st_mtime).strftime("%d.%m.%Y %H:%M"),
        }

        if HAS_WIN32:
            try:
                acad = get_autocad()
                doc  = deschide_dwg(acad, dwg_path)

                # Numar entitati
                info["nr_entitati"] = doc.ModelSpace.Count

                # Layere
                layere = [l.Name for l in doc.Layers]
                info["nr_layere"]   = len(layere)
                info["layere"]      = ", ".join(sorted(layere)[:20])

                doc.Close(False)
            except:
                info["nr_entitati"] = "—"
                info["nr_layere"]   = "—"

        raport.append(info)
        print(f"  {dwg_path.name}: {info['dimensiune_kb']} KB")

    # Excel raport
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Audit DWG"

    from openpyxl.styles import Font, PatternFill, Alignment
    coloane = ["fisier", "cale", "dimensiune_kb", "ultima_modif", "nr_entitati", "nr_layere", "layere"]
    headers = ["Fisier DWG", "Cale", "Dimensiune (KB)", "Ultima modif.", "Nr. entitati", "Nr. layere", "Layere (primele 20)"]

    for col_idx, (col, header) in enumerate(zip(coloane, headers), 1):
        celula = ws.cell(row=1, column=col_idx, value=header)
        celula.fill = PatternFill("solid", fgColor="1F3864")
        celula.font = Font(bold=True, color="FFFFFF")

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["G"].width = 60

    for row_idx, info in enumerate(raport, 2):
        for col_idx, col in enumerate(coloane, 1):
            ws.cell(row=row_idx, column=col_idx, value=info.get(col, ""))

    if not fisier_raport:
        fisier_raport = f"Audit_DWG_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

    wb.save(fisier_raport)
    print(f"\nRaport audit: {fisier_raport}")
    return fisier_raport


# ─────────────────────────────────────────────────────
#  CLI
# ─────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Operatii batch pe fisiere DWG — Studio Office Kolectiv",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Exemple:
  python batch-operatii.py --actiune plot        --folder ./desene  --output ./pdf
  python batch-operatii.py --actiune extrage     --folder ./desene  --output raport.xlsx
  python batch-operatii.py --actiune layere-arhi --folder ./desene
  python batch-operatii.py --actiune layere-urb  --folder ./desene
  python batch-operatii.py --actiune audit       --folder ./desene  --output audit.xlsx
        """
    )
    parser.add_argument("--actiune", required=True,
                        choices=["plot", "extrage", "layere-arhi", "layere-urb", "audit"],
                        help="Actiunea de executat")
    parser.add_argument("--folder",  required=True, help="Folder cu fisierele DWG")
    parser.add_argument("--output",  default=None,  help="Fisier/folder de output")

    args = parser.parse_args()

    print("=" * 55)
    print("  Batch DWG — Studio Office Kolectiv")
    print("=" * 55)

    if args.actiune == "plot":
        export_pdf_batch(args.folder, args.output or "./pdf-export")

    elif args.actiune == "extrage":
        extrage_atribute_blocuri(args.folder, args.output)

    elif args.actiune == "layere-arhi":
        aplica_layere_standard(args.folder, tip="arhi")

    elif args.actiune == "layere-urb":
        aplica_layere_standard(args.folder, tip="urb")

    elif args.actiune == "audit":
        audit_fisiere(args.folder, args.output)


if __name__ == "__main__":
    main()
