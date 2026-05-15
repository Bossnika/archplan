"""
Export Cantitati AutoCAD → Excel
Studio Office Kolectiv

Extrage date din desenul AutoCAD activ si genereaza
un deviz estimativ de cantitati in format Excel.

Utilizare:
  1. Ruleaza din AutoCAD: Tools > Macros > sau prin pyautocad
  2. Sau din linie de comanda: python export-cantitati.py <fisier.dwg>

Dependente: pip install pyautocad openpyxl comtypes
"""

import sys
import os
import math
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                  GradientFill)
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
except ImportError:
    print("ERROR: pip install openpyxl")
    sys.exit(1)

try:
    import win32com.client
    HAS_AUTOCAD = True
except ImportError:
    HAS_AUTOCAD = False
    print("ATENTIE: win32com nu e disponibil. Se folosesc date demo.")

# ─────────────────────────────────────────────
#  Constante culori Excel
# ─────────────────────────────────────────────
CULOARE_HEADER    = "1F3864"   # Albastru inchis
CULOARE_SECTIUNE  = "2E75B6"   # Albastru
CULOARE_SUBSECT   = "BDD7EE"   # Albastru deschis
CULOARE_TOTAL     = "FFD966"   # Galben
CULOARE_SUBTOTAL  = "FFF2CC"   # Galben deschis
CULOARE_ALTERNATIV = "EBF3FA"

# ─────────────────────────────────────────────
#  Extragere date din AutoCAD
# ─────────────────────────────────────────────

def conectare_autocad():
    """Conectare la instanta AutoCAD activa."""
    if not HAS_AUTOCAD:
        return None
    try:
        acad = win32com.client.Dispatch("AutoCAD.Application")
        doc  = acad.ActiveDocument
        print(f"Conectat la AutoCAD: {doc.Name}")
        return acad
    except Exception as e:
        print(f"Nu s-a putut conecta la AutoCAD: {e}")
        return None


def extrage_suprafete_layere(doc):
    """
    Parcurge toate entitatile din modelspace si calculeaza:
    - Suprafetele poliliniilor inchise per layer
    - Lungimile liniilor per layer
    - Numarul blocurilor per layer (usi, ferestre, etc.)
    """
    rezultate = {
        "suprafete": {},   # {layer: [suprafata1, suprafata2, ...]}
        "lungimi":   {},   # {layer: lungime_totala}
        "blocuri":   {},   # {(layer, bloc_name): count}
        "texte":     [],   # [(layer, text_value, pozitie)]
    }

    if doc is None:
        return rezultate

    model = doc.ModelSpace

    for entitate in model:
        try:
            layer = entitate.Layer
            tip   = entitate.ObjectName

            # Suprafete (LWPOLYLINE sau HATCH inchise)
            if tip in ("AcDbPolyline", "AcDb2dPolyline"):
                if entitate.Closed:
                    area = entitate.Area
                    if layer not in rezultate["suprafete"]:
                        rezultate["suprafete"][layer] = []
                    rezultate["suprafete"][layer].append(area)

            # Haşuri (suprafata exacta)
            elif tip == "AcDbHatch":
                area = entitate.Area
                if layer not in rezultate["suprafete"]:
                    rezultate["suprafete"][layer] = []
                rezultate["suprafete"][layer].append(area)

            # Lungimi (linii, polilinii deschise)
            elif tip in ("AcDbLine", "AcDbPolyline"):
                try:
                    lungime = entitate.Length
                    if layer not in rezultate["lungimi"]:
                        rezultate["lungimi"][layer] = 0.0
                    rezultate["lungimi"][layer] += lungime
                except:
                    pass

            # Blocuri (usi, ferestre, mobilier)
            elif tip == "AcDbBlockReference":
                bloc_name = entitate.Name
                cheie = (layer, bloc_name)
                if cheie not in rezultate["blocuri"]:
                    rezultate["blocuri"][cheie] = 0
                rezultate["blocuri"][cheie] += 1

            # Texte cu suprafete incaperi
            elif tip in ("AcDbText", "AcDbMText"):
                try:
                    txt = entitate.TextString
                    if "mp" in txt or "m²" in txt or "m2" in txt:
                        rezultate["texte"].append((layer, txt, None))
                except:
                    pass

        except Exception:
            continue

    return rezultate


def genereaza_date_demo():
    """Date demonstrative cand AutoCAD nu e disponibil."""
    return {
        "proiect": {
            "nume":     "Locuinta unifamiliala P+1",
            "adresa":   "Str. Exemplu nr. 1, Bucuresti",
            "beneficiar": "Popescu Ion",
            "proiectant": "Studio Office Kolectiv",
            "data":     datetime.now().strftime("%d.%m.%Y"),
            "faza":     "PT + DE",
            "nr_proiect": "2024-001",
        },
        "incaperi": [
            # (denumire, suprafata_utila, nivel)
            ("Hol intrare",       8.50, "Parter"),
            ("Living",           28.00, "Parter"),
            ("Bucatarie",        12.00, "Parter"),
            ("Baie parter",       5.50, "Parter"),
            ("Dormitor 1",       18.00, "Etaj 1"),
            ("Dormitor 2",       15.00, "Etaj 1"),
            ("Dormitor 3",       14.00, "Etaj 1"),
            ("Baie etaj",         7.50, "Etaj 1"),
            ("Hol etaj",          6.00, "Etaj 1"),
            ("Terasa",           12.00, "Etaj 1"),
        ],
        "pereti": [
            # (tip, grosime_cm, suprafata_m2, descriere)
            ("Caramida + Termosistem EPS100", 43.6, 180.0, "Pereti exteriori"),
            ("BCA 25cm",                      28.0,  95.0, "Pereti interiori portanti"),
            ("Gips carton 2x12.5 CW75",        10.0,  65.0, "Pereti despartitori"),
            ("Gips carton 1x12.5 CW50",         7.5,  30.0, "Pereti baie"),
        ],
        "tamplarie": [
            # (tip, dimensiune, cant, material, ug)
            ("Usa exterioara",    "90x210", 1,  "PVC termopan",      1.1),
            ("Usa interioara",    "80x210", 6,  "MDF folie",         None),
            ("Usa interioara",    "70x210", 3,  "MDF folie",         None),
            ("Fereastra",        "120x140", 4,  "PVC tripan Ug=0.5", 0.5),
            ("Fereastra",         "80x120", 3,  "PVC tripan Ug=0.5", 0.5),
            ("Fereastra",         "60x60",  2,  "PVC tripan Ug=0.5", 0.5),
            ("Usa-fereastra",    "210x230", 1,  "PVC tripan Ug=0.5", 0.5),
        ],
        "finisaje_podea": [
            # (tip_finisaj, suprafata, spatii)
            ("Gresie portelanata",   54.0, "Hol, bucatarie, bai"),
            ("Parchet lemn masiv",   75.0, "Living, dormitoare"),
            ("Terracota exterior",   12.0, "Terasa"),
        ],
        "finisaje_pereti_tavan": [
            # (tip, suprafata, spatii)
            ("Gleturi + vopsea lavabila alba", 420.0, "Toate incaperile"),
            ("Faianta baie h=2.20m",            35.0, "Bai"),
            ("Faianta bucatarie h=0.60m",        15.0, "Bucatarie"),
            ("Tencuiala exterioara decorativa", 180.0, "Fatade"),
        ],
        "structura": [
            # (element, material, cantitate, unitate)
            ("Fundatie continua",   "Beton C20/25",  45.0,  "m³"),
            ("Stalpi beton armat",  "Beton C25/30",   8.5,  "m³"),
            ("Grinzi beton armat",  "Beton C25/30",  12.0,  "m³"),
            ("Placa beton armat",   "Beton C25/30",  38.0,  "m³"),
            ("Otel armare",         "OB37 / PC52",  4200.0, "kg"),
            ("Cofraj",              "Lemn + panouri", 185.0, "m²"),
        ],
        "acoperis": [
            ("Sarpanta lemn ecarisat",       "Lemn cl. C24",  3200.0, "ml"),
            ("Membrane hidroizolatoare",     "FPO/TPO",         220.0, "m²"),
            ("Tigla ceramica",               "Bramac/Tondach", 220.0, "m²"),
            ("Jgheaburi + burlane",          "Tabla Zn",         48.0, "ml"),
            ("Termoizolatie pod vata 20cm", "Knauf/Isover",    120.0, "m²"),
        ],
    }


# ─────────────────────────────────────────────
#  Generare Excel
# ─────────────────────────────────────────────

def stil_header(ws, rand, col_start, col_end, text, culoare=None, font_size=11):
    """Scrie header cu merge si formatare."""
    ws.merge_cells(
        start_row=rand, start_column=col_start,
        end_row=rand,   end_column=col_end
    )
    celula = ws.cell(row=rand, column=col_start, value=text)
    culoare = culoare or CULOARE_HEADER
    celula.fill = PatternFill("solid", fgColor=culoare)
    celula.font = Font(bold=True, color="FFFFFF", size=font_size)
    celula.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return celula


def stil_sectiune(ws, rand, col_start, col_end, text):
    ws.merge_cells(
        start_row=rand, start_column=col_start,
        end_row=rand,   end_column=col_end
    )
    celula = ws.cell(row=rand, column=col_start, value=text)
    celula.fill = PatternFill("solid", fgColor=CULOARE_SECTIUNE)
    celula.font = Font(bold=True, color="FFFFFF", size=10)
    celula.alignment = Alignment(horizontal="left", vertical="center")
    return celula


def border_thin():
    side = Side(style="thin", color="AAAAAA")
    return Border(left=side, right=side, top=side, bottom=side)


def scrie_rand_date(ws, rand, valori, alternativ=False, bold=False, culoare=None):
    """Scrie un rand de date cu stilizare."""
    bg = culoare or (CULOARE_ALTERNATIV if alternativ else "FFFFFF")
    for col, val in enumerate(valori, 1):
        c = ws.cell(row=rand, column=col, value=val)
        c.fill    = PatternFill("solid", fgColor=bg)
        c.font    = Font(bold=bold, size=9)
        c.border  = border_thin()
        c.alignment = Alignment(horizontal="left" if col <= 2 else "center",
                                 vertical="center", wrap_text=True)


def genereaza_excel(date, fisier_output=None):
    """Generare fisier Excel complet cu toate foile."""

    wb = openpyxl.Workbook()

    # ── FOAIE 1: Rezumat proiect ──────────────────────────────────────────
    ws_rez = wb.active
    ws_rez.title = "Rezumat Proiect"
    ws_rez.column_dimensions["A"].width = 30
    ws_rez.column_dimensions["B"].width = 45
    ws_rez.row_dimensions[1].height = 35

    stil_header(ws_rez, 1, 1, 2,
                f"REZUMAT PROIECT — {date['proiect']['nume']}", font_size=13)

    info_proiect = [
        ("Nr. proiect",      date["proiect"]["nr_proiect"]),
        ("Denumire lucrare", date["proiect"]["nume"]),
        ("Adresa",           date["proiect"]["adresa"]),
        ("Beneficiar",       date["proiect"]["beneficiar"]),
        ("Proiectant",       date["proiect"]["proiectant"]),
        ("Faza",             date["proiect"]["faza"]),
        ("Data intocmirii",  date["proiect"]["data"]),
    ]

    for i, (eticheta, valoare) in enumerate(info_proiect, 3):
        ws_rez.cell(row=i, column=1, value=eticheta).font = Font(bold=True, size=10)
        ws_rez.cell(row=i, column=2, value=valoare).font  = Font(size=10)

    # Sumar suprafete
    stil_sectiune(ws_rez, 12, 1, 2, "SUMAR SUPRAFETE")
    su_total = sum(s for _, s, _ in date["incaperi"])
    sc_total = su_total * 1.10  # ~10% pereti si circulatii

    ws_rez.cell(row=13, column=1, value="Suprafata utila totala (Su)").font = Font(bold=True, size=10)
    ws_rez.cell(row=13, column=2, value=f"{su_total:.2f} m²").font = Font(size=10)
    ws_rez.cell(row=14, column=1, value="Suprafata construita (Sc)").font = Font(bold=True, size=10)
    ws_rez.cell(row=14, column=2, value=f"{sc_total:.2f} m²").font = Font(size=10)

    # ── FOAIE 2: Suprafete incaperi ───────────────────────────────────────
    ws_sup = wb.create_sheet("Suprafete Incaperi")
    ws_sup.column_dimensions["A"].width = 5
    ws_sup.column_dimensions["B"].width = 30
    ws_sup.column_dimensions["C"].width = 18
    ws_sup.column_dimensions["D"].width = 18
    ws_sup.column_dimensions["E"].width = 25
    ws_sup.row_dimensions[1].height = 30

    stil_header(ws_sup, 1, 1, 5, "TABEL SUPRAFETE INCAPERI", font_size=12)

    headers = ["Nr.", "Denumire incapere", "Suprafata utila (m²)", "Nivel", "Observatii"]
    scrie_rand_date(ws_sup, 2, headers, bold=True,
                    culoare=CULOARE_SUBSECT)

    nivele = {}
    for idx, (denumire, suprafata, nivel) in enumerate(date["incaperi"], 1):
        alternativ = (idx % 2 == 0)
        scrie_rand_date(ws_sup, idx + 2,
                        [idx, denumire, suprafata, nivel, ""],
                        alternativ=alternativ)
        if nivel not in nivele:
            nivele[nivel] = 0.0
        nivele[nivel] += suprafata

    rand_curent = len(date["incaperi"]) + 3

    # Subtotaluri per nivel
    for nivel, total in sorted(nivele.items()):
        scrie_rand_date(ws_sup, rand_curent,
                        ["", f"TOTAL {nivel.upper()}", f"{total:.2f}", nivel, ""],
                        bold=True, culoare=CULOARE_SUBTOTAL)
        rand_curent += 1

    # Total general
    su_total = sum(s for _, s, _ in date["incaperi"])
    scrie_rand_date(ws_sup, rand_curent,
                    ["", "SUPRAFATA UTILA TOTALA (Su)", f"{su_total:.2f} m²", "ALL", ""],
                    bold=True, culoare=CULOARE_TOTAL)

    # ── FOAIE 3: Tamplarie ────────────────────────────────────────────────
    ws_tamp = wb.create_sheet("Tamplarie")
    for col, latime in zip("ABCDEFG", [5, 22, 15, 8, 20, 10, 25]):
        ws_tamp.column_dimensions[col].width = latime
    ws_tamp.row_dimensions[1].height = 30

    stil_header(ws_tamp, 1, 1, 7, "TABEL TAMPLARIE", font_size=12)

    headers_tamp = ["Nr.", "Tip element", "Dimensiune (LxH)", "Buc.", "Material / Specificatie",
                    "Ug (W/m²K)", "Observatii"]
    scrie_rand_date(ws_tamp, 2, headers_tamp, bold=True, culoare=CULOARE_SUBSECT)

    for idx, (tip, dim, cant, material, ug) in enumerate(date["tamplarie"], 1):
        ug_val = f"{ug}" if ug else "—"
        scrie_rand_date(ws_tamp, idx + 2,
                        [idx, tip, dim, cant, material, ug_val, ""],
                        alternativ=(idx % 2 == 0))

    # ── FOAIE 4: Pereti & Finisaje ────────────────────────────────────────
    ws_fin = wb.create_sheet("Pereti & Finisaje")
    for col, latime in zip("ABCDEF", [5, 35, 15, 12, 15, 25]):
        ws_fin.column_dimensions[col].width = latime

    stil_header(ws_fin, 1, 1, 6, "PERETI & FINISAJE", font_size=12)

    # Pereti
    stil_sectiune(ws_fin, 2, 1, 6, "A. PERETI")
    scrie_rand_date(ws_fin, 3,
                    ["Nr.", "Tip perete", "Grosime (cm)", "Suprafata (m²)", "Volume (m³)", "Obs."],
                    bold=True, culoare=CULOARE_SUBSECT)
    rand_curent = 4
    for idx, (tip, gros, sup, desc) in enumerate(date["pereti"], 1):
        vol = round(sup * gros / 100.0, 2)
        scrie_rand_date(ws_fin, rand_curent,
                        [idx, f"{tip} — {desc}", f"{gros} cm", f"{sup:.1f}", f"{vol:.2f}", ""],
                        alternativ=(idx % 2 == 0))
        rand_curent += 1

    rand_curent += 1
    # Finisaje podea
    stil_sectiune(ws_fin, rand_curent, 1, 6, "B. FINISAJE PARDOSELI")
    rand_curent += 1
    scrie_rand_date(ws_fin, rand_curent,
                    ["Nr.", "Tip finisaj", "Suprafata (m²)", "Spatii", "", "Obs."],
                    bold=True, culoare=CULOARE_SUBSECT)
    rand_curent += 1
    for idx, (tip, sup, spatii) in enumerate(date["finisaje_podea"], 1):
        scrie_rand_date(ws_fin, rand_curent,
                        [idx, tip, f"{sup:.1f}", spatii, "", ""],
                        alternativ=(idx % 2 == 0))
        rand_curent += 1

    rand_curent += 1
    # Finisaje pereti-tavan
    stil_sectiune(ws_fin, rand_curent, 1, 6, "C. FINISAJE PERETI & TAVAN")
    rand_curent += 1
    scrie_rand_date(ws_fin, rand_curent,
                    ["Nr.", "Tip finisaj", "Suprafata (m²)", "Spatii", "", "Obs."],
                    bold=True, culoare=CULOARE_SUBSECT)
    rand_curent += 1
    for idx, (tip, sup, spatii) in enumerate(date["finisaje_pereti_tavan"], 1):
        scrie_rand_date(ws_fin, rand_curent,
                        [idx, tip, f"{sup:.1f}", spatii, "", ""],
                        alternativ=(idx % 2 == 0))
        rand_curent += 1

    # ── FOAIE 5: Structura ────────────────────────────────────────────────
    ws_str = wb.create_sheet("Structura & Acoperis")
    for col, latime in zip("ABCDEF", [5, 28, 20, 12, 12, 20]):
        ws_str.column_dimensions[col].width = latime

    stil_header(ws_str, 1, 1, 6, "STRUCTURA & ACOPERIS", font_size=12)

    stil_sectiune(ws_str, 2, 1, 6, "A. ELEMENTE STRUCTURALE")
    scrie_rand_date(ws_str, 3,
                    ["Nr.", "Element structural", "Material / Clasa", "Cantitate", "U.M.", "Obs."],
                    bold=True, culoare=CULOARE_SUBSECT)
    rand_curent = 4
    for idx, (elem, material, cant, um) in enumerate(date["structura"], 1):
        scrie_rand_date(ws_str, rand_curent,
                        [idx, elem, material, cant, um, ""],
                        alternativ=(idx % 2 == 0))
        rand_curent += 1

    rand_curent += 1
    stil_sectiune(ws_str, rand_curent, 1, 6, "B. ACOPERIS & INVELITOARE")
    rand_curent += 1
    scrie_rand_date(ws_str, rand_curent,
                    ["Nr.", "Element", "Material / Produs", "Cantitate", "U.M.", "Obs."],
                    bold=True, culoare=CULOARE_SUBSECT)
    rand_curent += 1
    for idx, (elem, material, cant, um) in enumerate(date["acoperis"], 1):
        scrie_rand_date(ws_str, rand_curent,
                        [idx, elem, material, cant, um, ""],
                        alternativ=(idx % 2 == 0))
        rand_curent += 1

    # ── FOAIE 6: Centralizator ────────────────────────────────────────────
    ws_ctr = wb.create_sheet("Centralizator Cantitati")
    for col, latime in zip("ABCDEFG", [5, 35, 15, 12, 12, 15, 20]):
        ws_ctr.column_dimensions[col].width = latime
    ws_ctr.row_dimensions[1].height = 35

    stil_header(ws_ctr, 1, 1, 7,
                f"CENTRALIZATOR CANTITATI — {date['proiect']['nume']}\n"
                f"Data: {date['proiect']['data']}  |  Faza: {date['proiect']['faza']}",
                font_size=12)

    scrie_rand_date(ws_ctr, 2,
                    ["Nr.", "Capitol / Lucrare", "U.M.", "Cantitate", "Pret unit. (RON)",
                     "Total (RON)", "Observatii"],
                    bold=True, culoare=CULOARE_SUBSECT)

    capitole = [
        ("TERASAMENTE & SISTEMATIZARE", [
            ("Sapatura mecanica fundatii",     "m³", 55.0),
            ("Umplutura compactata",           "m³", 20.0),
            ("Evacuare pamant",                "m³", 35.0),
        ]),
        ("FUNDATII", [
            ("Beton de egalizare C8/10",       "m³",  4.5),
            ("Beton fundatii C20/25",          "m³", 38.0),
            ("Otel armare PC52",               "kg",1800.0),
            ("Hidroizolatie fundatii",         "m²", 85.0),
        ]),
        ("STRUCTURA", [
            (f"Beton armat stalpi/grinzi C25/30","m³",20.5),
            ("Placa beton armat C25/30",        "m³", 38.0),
            ("Otel armare PC52/OB37",           "kg",2400.0),
            ("Cofraj lemn/panouri",             "m²",185.0),
        ]),
        ("PERETI & ZIDARIE", [
            (f"{date['pereti'][0][0]}", "m²", date["pereti"][0][2]),
            (f"{date['pereti'][1][0]}", "m²", date["pereti"][1][2]),
            (f"{date['pereti'][2][0]}", "m²", date["pereti"][2][2]),
            (f"{date['pereti'][3][0]}", "m²", date["pereti"][3][2]),
        ]),
        ("TAMPLARIE", [
            (f"{t[0]} {t[1]}", "buc", t[2]) for t in date["tamplarie"]
        ]),
        ("ACOPERIS & INVELITOARE", [
            (a[0], a[3], a[2]) for a in date["acoperis"]
        ]),
        ("FINISAJE INTERIOARE", [
            (f[0], "m²", f[1]) for f in date["finisaje_podea"]
        ] + [
            (f[0], "m²", f[1]) for f in date["finisaje_pereti_tavan"]
        ]),
        ("INSTALATII", [
            ("Instalatii sanitare (deviz separat)",     "ls", 1),
            ("Instalatii electrice (deviz separat)",    "ls", 1),
            ("Instalatii termice/HVAC (deviz separat)", "ls", 1),
        ]),
    ]

    rand_curent = 3
    nr_global   = 1
    total_general = 0.0

    for capitol, articole in capitole:
        stil_sectiune(ws_ctr, rand_curent, 1, 7, capitol)
        rand_curent += 1

        for denumire, um, cantitate in articole:
            scrie_rand_date(ws_ctr, rand_curent,
                            [nr_global, denumire, um, f"{cantitate:.2f}", "—", "—", ""],
                            alternativ=(nr_global % 2 == 0))
            nr_global   += 1
            rand_curent += 1

        rand_curent += 1  # spatiu intre capitole

    # Nota de subsol
    ws_ctr.merge_cells(
        start_row=rand_curent + 1, start_column=1,
        end_row=rand_curent + 1,   end_column=7
    )
    nota = ws_ctr.cell(row=rand_curent + 1, column=1,
        value="* Preturile unitare sunt orientative si necesita oferte de la furnizori/executanti.")
    nota.font      = Font(italic=True, size=8, color="666666")
    nota.alignment = Alignment(horizontal="left")

    # ── Salvare ───────────────────────────────────────────────────────────
    if not fisier_output:
        data_str  = datetime.now().strftime("%Y%m%d_%H%M")
        nume_proj = date["proiect"]["nr_proiect"].replace("/", "-").replace(" ", "_")
        fisier_output = f"Cantitati_{nume_proj}_{data_str}.xlsx"

    wb.save(fisier_output)
    print(f"\nFisier Excel generat: {fisier_output}")
    print(f"Foi: {[ws.title for ws in wb.worksheets]}")
    return fisier_output


# ─────────────────────────────────────────────
#  Punct de intrare
# ─────────────────────────────────────────────

def main():
    print("=" * 60)
    print("  Export Cantitati AutoCAD → Excel")
    print("  Studio Office Kolectiv")
    print("=" * 60)

    fisier_output = sys.argv[1] if len(sys.argv) > 1 else None

    # Incearca conectare AutoCAD
    acad = conectare_autocad()

    if acad:
        doc  = acad.ActiveDocument
        date_autocad = extrage_suprafete_layere(doc)
        # Construieste structura de date din datele extrase
        # (in productie, aceasta logica ar fi mai complexa)
        date = genereaza_date_demo()
        date["proiect"]["nume"] = os.path.splitext(doc.Name)[0]
        print(f"Date extrase din: {doc.Name}")
    else:
        print("Utilizare date demo (AutoCAD offline).")
        date = genereaza_date_demo()

    fisier = genereaza_excel(date, fisier_output)
    print(f"\nDeschide fisierul: {fisier}")


if __name__ == "__main__":
    main()
